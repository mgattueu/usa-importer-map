import openpyxl, re
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

wb_in = openpyxl.load_workbook(r'C:\Users\mgatt\Downloads\ImportYeti-PowerQuery-05-16-2026.xlsx', read_only=True)
ws_in = wb_in.active
rows = list(ws_in.iter_rows(values_only=True))

def extract_state(addr):
    if not addr: return 'UNKNOWN'
    m = re.findall(r'\b([A-Z]{2})\s*\d{5}\b', str(addr).upper())
    return m[0] if m else 'UNKNOWN'

def extract_city(addr):
    if not addr: return ''
    addr_str = str(addr)
    first = addr_str.split(',')[0].strip()
    m = re.search(r'([A-Za-z ]+?)\s+([A-Z]{2})\s*\d{5}', first)
    if m:
        return m.group(1).strip().title()
    return ''

def extract_phone(addr):
    if not addr: return ''
    m = re.search(r'(?:Tel|Phone|Ph)[:\s]*([0-9\-\(\)\s]{7,20})', str(addr), re.IGNORECASE)
    return m.group(1).strip() if m else ''

def extract_email(addr):
    if not addr: return ''
    m = re.search(r'([a-zA-Z0-9._%+-]+\s*@\s*[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})', str(addr))
    return m.group(1).replace(' ', '').lower() if m else ''

LOGISTICS_KEYWORDS = [
    'logistics','freight','maritime','customs broker','forwarding','airlift',
    'interglobo','oec ','oec freight','pegasus maritime','firstlift','patagon',
    'comfy','brightway','ctc logistics','relay logistics','worldwide logistics',
    'imperative logistics','apex maritime','cil freight','umax shipping',
    'interfreight','leela logistics','mts logistics','blackstone shipping',
    'ups ocean','penguin shipping','worldwide logistics partners',
    'american global freight','oocl','c&l container','rtw logistics',
    'jenson logistics','topocean','city ocean','transmarine','em lines',
    'famous pacific','pan pacific','ctl lax','car go worldwide','amass global',
    'affinity shipping','cts global supply','dsv air','cms shipping',
    'western overseas','ascend express','sea dominion','binex line',
    'orient express container','crossea shipping','seahorse container',
    'seax trade','seko worldwide','hecny','winfar','stable enterprise',
    'permeco','best global management','tanera transport','jdy international',
    'db group america','jupiter international','d b group','general noli',
    'ftl plus','world class shipping','skytrans','jr global','del corona',
    'gh trans','speedier logistic','oriental air','advantage group intl',
    'shipco','comage container','glenrock international','a j worldwide',
    '1up cargo','gran trade','express consolidation','aetos cargo','vg enterprises',
    'tlss inc','aprile usa','troy container','city ocean intl','magellan shipping',
    'gateway international','norman krieger','john s connor','unitrans',
    'ch robinson','alliance trade','tql global','midwest transatlantic',
    'columbus customhouse','sjlt usa','triumph express','shipco transport',
    'traffic tech','europa ltd','bbe expediting','ecu worldwide','acrocargo',
    'master logistix','zehong global','harvest logistic','cole international',
    'canaan transport','booking union','ultra air cargo','savino del bene',
    'sarlogisolutions','db shipping','vnft international','rainbow import',
    'fcg global','posey international','associated import','carnevale',
    'locher evers','seair global','dsv','rohlig','shipco',
]

def categorize(name, product):
    n = name.lower()
    p = (product or '').lower()
    if 'kani international' in n: return 'SELF'
    if 'slab planet' in n: return 'SELF'
    for k in LOGISTICS_KEYWORDS:
        if k in n: return 'LOGISTICS'
    # Known competitors
    if any(x in n for x in ['quarr','granite','stone','monument','marble','memorial',
                              'nustone','cosentino','mausoleum','cemetery supply',
                              'headstone','tombstone']):
        return 'COMPETITOR'
    if any(x in n for x in ['york group','blast craft','harris enterprises','buchanan group',
                              'boc international','kerry apex','sabbow','alpi usa',
                              'matthews international','ots astracon','star asia',
                              'great wall stone','wds granite','stone quest',
                              'olympiad line','tecstone','eagle granite','alpi']):
        return 'COMPETITOR'
    # Potential buyers
    if any(x in n for x in ['cemetery','funeral','cremation','mausoleum',
                              'memorial park','burial','casket','mortuary']):
        return 'BUYER'
    return 'INVESTIGATE'

REGIONS = {
    'Northeast': ['NJ','NY','MA','CT','RI','VT','NH','ME','PA','MD','DE'],
    'Southeast': ['GA','NC','SC','FL','VA','AL','MS','TN','KY','LA','AR'],
    'Midwest':   ['IL','OH','IN','MI','WI','MN','IA','MO','KS'],
    'South-TX':  ['TX','OK'],
    'West':      ['CA','WA','OR','NV','UT','CO','AZ','ID','MT','WY'],
}
state_to_region = {s: r for r, states in REGIONS.items() for s in states}

data = []
for r in rows[1:]:
    vol = r[0] or 0
    teu = r[1] or 0
    name = r[2] or ''
    product = r[3] or ''
    notify_addr = r[4] or ''
    state = extract_state(notify_addr)
    city = extract_city(notify_addr)
    cat = categorize(name, product)
    phone = extract_phone(notify_addr)
    email = extract_email(notify_addr)
    region = state_to_region.get(state, 'Other/International')
    data.append({
        'vol': vol, 'teu': teu, 'name': name, 'state': state,
        'city': city, 'cat': cat, 'phone': phone, 'email': email,
        'region': region, 'addr': str(notify_addr)[:120]
    })

# ── Build output workbook ────────────────────────────────────────────
wb = openpyxl.Workbook()

# ── Color palette ────────────────────────────────────────────────────
RED    = PatternFill('solid', fgColor='FFCCCC')
GREEN  = PatternFill('solid', fgColor='C6EFCE')
YELLOW = PatternFill('solid', fgColor='FFFACD')
BLUE   = PatternFill('solid', fgColor='BDD7EE')
GRAY   = PatternFill('solid', fgColor='EEEEEE')
ORANGE = PatternFill('solid', fgColor='FFE4B5')
HDR    = PatternFill('solid', fgColor='1F4E79')
hdr_font = Font(bold=True, color='FFFFFF', size=11)
bold = Font(bold=True)

thin = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def set_header(ws, cols):
    for i, h in enumerate(cols, 1):
        c = ws.cell(1, i, h)
        c.fill = HDR
        c.font = hdr_font
        c.alignment = Alignment(horizontal='center', wrap_text=True)
        c.border = thin

def col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def freeze(ws):
    ws.freeze_panes = 'A2'

# ══════════════════════════════════════════════════
# SHEET 1: VISIT PLAN (competitors=red, buyers=green, investigate=yellow)
# ══════════════════════════════════════════════════
ws1 = wb.active
ws1.title = '🗺 Visit Plan'

cats_order = ['BUYER','INVESTIGATE','COMPETITOR','LOGISTICS','SELF']
cat_fill = {'BUYER': GREEN, 'INVESTIGATE': YELLOW, 'COMPETITOR': RED,
            'LOGISTICS': GRAY, 'SELF': BLUE}
cat_label = {'BUYER': 'VISIT - End Buyer (cemetery/funeral)',
             'INVESTIGATE': 'VISIT - Investigate (dealer?)',
             'COMPETITOR': 'SKIP - Direct Competitor',
             'LOGISTICS': 'SKIP - Freight/Logistics only',
             'SELF': 'SELF'}

cols = ['Action', 'Company Name', 'State', 'City', 'Volume (TEU)', 'Category', 'Phone', 'Email', 'Region', 'Address']
set_header(ws1, cols)

sorted_data = sorted(data, key=lambda x: (cats_order.index(x['cat']), x['state'], -x['vol']))
for i, d in enumerate(sorted_data, 2):
    action = cat_label[d['cat']]
    row = [action, d['name'], d['state'], d['city'], d['vol'],
           d['cat'], d['phone'], d['email'], d['region'], d['addr']]
    fill = cat_fill[d['cat']]
    for j, val in enumerate(row, 1):
        c = ws1.cell(i, j, val)
        c.fill = fill
        c.border = thin
        c.alignment = Alignment(wrap_text=False)

col_widths(ws1, [30, 42, 8, 20, 14, 18, 18, 28, 18, 60])
ws1.auto_filter.ref = f'A1:{get_column_letter(len(cols))}1'
freeze(ws1)

# ══════════════════════════════════════════════════
# SHEET 2: HOT LEADS (buyers + investigate, vol >= 5)
# ══════════════════════════════════════════════════
ws2 = wb.create_sheet('🎯 Hot Leads')
cols2 = ['Priority', 'Company Name', 'State', 'City', 'Volume', 'Why Visit', 'Phone', 'Email', 'Address']
set_header(ws2, cols2)

hot = [d for d in data if d['cat'] in ('BUYER', 'INVESTIGATE') and d['vol'] >= 5]
hot.sort(key=lambda x: -x['vol'])
for i, d in enumerate(hot, 2):
    priority = 'HIGH' if d['vol'] >= 20 else ('MED' if d['vol'] >= 10 else 'LOW')
    why = 'Cemetery/Funeral - end buyer' if d['cat'] == 'BUYER' else 'Potential dealer/distributor'
    row = [priority, d['name'], d['state'], d['city'], d['vol'], why,
           d['phone'], d['email'], d['addr']]
    fill = GREEN if d['cat'] == 'BUYER' else YELLOW
    for j, val in enumerate(row, 1):
        c = ws2.cell(i, j, val)
        c.fill = fill
        c.border = thin

col_widths(ws2, [10, 42, 8, 20, 10, 34, 18, 28, 60])
ws2.auto_filter.ref = f'A1:{get_column_letter(len(cols2))}1'
freeze(ws2)

# ══════════════════════════════════════════════════
# SHEET 3: COMPETITORS - know thy enemy
# ══════════════════════════════════════════════════
ws3 = wb.create_sheet('⚠ Competitors')
cols3 = ['Company Name', 'State', 'City', 'Volume', 'Est. Market Share %', 'Region']
set_header(ws3, cols3)

total_comp_vol = sum(d['vol'] for d in data if d['cat'] == 'COMPETITOR')
comps = [d for d in data if d['cat'] == 'COMPETITOR']
comps.sort(key=lambda x: -x['vol'])
for i, d in enumerate(comps, 2):
    share = f"{d['vol'] / total_comp_vol * 100:.1f}%"
    row = [d['name'], d['state'], d['city'], d['vol'], share, d['region']]
    for j, val in enumerate(row, 1):
        c = ws3.cell(i, j, val)
        c.fill = RED
        c.border = thin

col_widths(ws3, [42, 8, 20, 10, 18, 18])
ws3.auto_filter.ref = f'A1:{get_column_letter(len(cols3))}1'
freeze(ws3)

# ══════════════════════════════════════════════════
# SHEET 4: WAREHOUSE EXPANSION MAP
# ══════════════════════════════════════════════════
ws4 = wb.create_sheet('🏭 Warehouse Expansion')

reg_data = defaultdict(lambda: {'comp':0,'buyer':0,'invest':0,'logistics':0,'comp_cnt':0,'buyer_cnt':0,'invest_cnt':0})
for d in data:
    if d['cat'] == 'SELF': continue
    reg = d['region']
    reg_data[reg]['comp'] += d['vol'] if d['cat']=='COMPETITOR' else 0
    reg_data[reg]['buyer'] += d['vol'] if d['cat']=='BUYER' else 0
    reg_data[reg]['invest'] += d['vol'] if d['cat']=='INVESTIGATE' else 0
    reg_data[reg]['logistics'] += d['vol'] if d['cat']=='LOGISTICS' else 0
    reg_data[reg]['comp_cnt'] += 1 if d['cat']=='COMPETITOR' else 0
    reg_data[reg]['buyer_cnt'] += 1 if d['cat']=='BUYER' else 0
    reg_data[reg]['invest_cnt'] += 1 if d['cat']=='INVESTIGATE' else 0

# Title
ws4['A1'] = 'WAREHOUSE EXPANSION OPPORTUNITY ANALYSIS'
ws4['A1'].font = Font(bold=True, size=14, color='1F4E79')
ws4['A2'] = 'Higher Opportunity % = less competitor coverage relative to total market demand'
ws4['A2'].font = Font(italic=True, color='666666')

cols4 = ['Region', 'Total Demand\n(all importers)', 'Competitor\nVolume',
         '# Competitors', 'Potential\nCustomers Vol', '# Potential\nCustomers',
         'Opportunity\n%', 'Recommendation', 'Best City to Place Warehouse']
for i, h in enumerate(cols4, 1):
    c = ws4.cell(4, i, h)
    c.fill = HDR
    c.font = hdr_font
    c.alignment = Alignment(horizontal='center', wrap_text=True)
    c.border = thin
ws4.row_dimensions[4].height = 40

WAREHOUSE_RECS = {
    'Northeast': ('Already served from Hillsborough NJ', 'Hillsborough, NJ (current)'),
    'Southeast': ('HIGH PRIORITY - Best opportunity ratio', 'Charlotte NC  -OR-  Atlanta GA'),
    'Midwest':   ('MEDIUM - High comp coverage, still worth serving', 'Columbus OH  -OR-  Chicago IL'),
    'South-TX':  ('MEDIUM - Growing market', 'Dallas-Fort Worth TX'),
    'West':      ('CONSIDER - Huge market, distant', 'Los Angeles CA  -OR-  Seattle WA'),
    'Other/International': ('Canada/Mexico/International - N/A', 'N/A'),
}

region_order = ['Northeast','Southeast','Midwest','South-TX','West','Other/International']
for i, reg in enumerate(region_order, 5):
    r = reg_data[reg]
    total = r['comp'] + r['buyer'] + r['invest'] + r['logistics']
    cust = r['buyer'] + r['invest']
    opp_pct = (cust / r['comp'] * 100) if r['comp'] > 0 else 999
    rec, city = WAREHOUSE_RECS.get(reg, ('', ''))

    fill = GREEN if opp_pct > 100 else (YELLOW if opp_pct > 40 else ORANGE)
    row = [reg, total, r['comp'], r['comp_cnt'], cust,
           r['buyer_cnt']+r['invest_cnt'], f"{min(opp_pct,999):.0f}%", rec, city]
    for j, val in enumerate(row, 1):
        c = ws4.cell(i, j, val)
        c.fill = fill
        c.border = thin
        c.alignment = Alignment(wrap_text=True, horizontal='center' if j > 1 else 'left')

col_widths(ws4, [18, 16, 16, 14, 18, 16, 14, 38, 32])
ws4.row_dimensions[5].height = 20
ws4.row_dimensions[6].height = 20

# Legend
ws4.cell(12, 1, 'LEGEND:').font = bold
ws4.cell(13, 1, 'Green = Best opportunity (cust vol > comp vol)').fill = GREEN
ws4.cell(14, 1, 'Yellow = Moderate opportunity').fill = YELLOW
ws4.cell(15, 1, 'Orange = Competitor-heavy (harder to enter)').fill = ORANGE

# ══════════════════════════════════════════════════
# SHEET 5: STATE BREAKDOWN
# ══════════════════════════════════════════════════
ws5 = wb.create_sheet('📊 By State')
cols5 = ['State', 'Region', 'Total Volume', '# Competitors', 'Competitor Vol',
         '# Potential Customers', 'Customer Vol', 'Opportunity Score']
set_header(ws5, cols5)

state_agg = defaultdict(lambda: {'reg':'','comp':0,'cust':0,'comp_cnt':0,'cust_cnt':0})
for d in data:
    if d['cat'] == 'SELF': continue
    s = d['state']
    state_agg[s]['reg'] = d['region']
    if d['cat'] == 'COMPETITOR':
        state_agg[s]['comp'] += d['vol']
        state_agg[s]['comp_cnt'] += 1
    elif d['cat'] in ('BUYER','INVESTIGATE'):
        state_agg[s]['cust'] += d['vol']
        state_agg[s]['cust_cnt'] += 1

state_rows = [(s, v) for s, v in state_agg.items() if s not in ('UNKNOWN','TE','HH','CP','NA','US')]
state_rows.sort(key=lambda x: -(x[1]['comp']+x[1]['cust']))

for i, (s, v) in enumerate(state_rows, 2):
    total = v['comp'] + v['cust']
    score = v['cust'] / v['comp'] * 100 if v['comp'] > 0 else 999
    row = [s, v['reg'], total, v['comp_cnt'], v['comp'], v['cust_cnt'], v['cust'], f"{min(score,999):.0f}%"]
    fill = GREEN if score > 80 else (YELLOW if score > 30 else RED)
    for j, val in enumerate(row, 1):
        c = ws5.cell(i, j, val)
        c.fill = fill
        c.border = thin

col_widths(ws5, [8, 18, 14, 16, 14, 20, 14, 18])
ws5.auto_filter.ref = f'A1:{get_column_letter(len(cols5))}1'
freeze(ws5)

out_path = r'C:\Users\mgatt\Downloads\Kani_VisitPlan_May2026.xlsx'
wb.save(out_path)
print(f"Saved: {out_path}")
print(f"\nSummary:")
print(f"  Total companies in data: {len(data)}")
print(f"  SKIP - Competitors: {sum(1 for d in data if d['cat']=='COMPETITOR')}")
print(f"  SKIP - Logistics:   {sum(1 for d in data if d['cat']=='LOGISTICS')}")
print(f"  VISIT - Buyers:     {sum(1 for d in data if d['cat']=='BUYER')}")
print(f"  VISIT - Investigate:{sum(1 for d in data if d['cat']=='INVESTIGATE')}")
print(f"  Self:               {sum(1 for d in data if d['cat']=='SELF')}")
print(f"\nHot Leads (vol >= 5): {sum(1 for d in data if d['cat'] in ('BUYER','INVESTIGATE') and d['vol']>=5)}")
